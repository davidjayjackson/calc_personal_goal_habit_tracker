"""
Build personal-goal-habit-tracker.ots
Generates an xlsx with sample data + charts, then converts to ots via LibreOffice.
"""

import subprocess
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Palette ────────────────────────────────────────────────────────────────────
INPUT_FILL   = PatternFill("solid", fgColor="D6EAF8")   # light blue  – user input
HEADER_FILL  = PatternFill("solid", fgColor="1A5276")   # dark blue   – headers
GREEN_FILL   = PatternFill("solid", fgColor="D5F5E3")   # light green – on track / done
RED_FILL     = PatternFill("solid", fgColor="FADBD8")   # light red   – missed / at risk
AMBER_FILL   = PatternFill("solid", fgColor="FDEBD0")   # amber       – warning
BLUE_FILL    = PatternFill("solid", fgColor="D6EAF8")   # blue        – achieved
GREY_FILL    = PatternFill("solid", fgColor="EAECEE")   # grey        – paused
DARK_GREEN   = PatternFill("solid", fgColor="1E8449")   # dark green  – streak
MID_GREEN    = PatternFill("solid", fgColor="82E0AA")   # mid green
LIGHT_GREEN  = PatternFill("solid", fgColor="D5F5E3")   # faint green
SUBHDR_FILL  = PatternFill("solid", fgColor="2E86C1")   # mid blue    – sub-headers

WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DARK_FONT   = Font(name="Calibri", bold=True, color="1A5276", size=11)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
BODY_FONT   = Font(name="Calibri", size=11)
BOLD_FONT   = Font(name="Calibri", bold=True, size=11)
SMALL_FONT  = Font(name="Calibri", size=9, italic=True, color="7F8C8D")

THIN = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def hdr(ws, row, col, value, fill=None, font=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill    = fill  or HEADER_FILL
    c.font    = font  or HEADER_FONT
    c.alignment = align or CENTER
    if border:
        c.border = BORDER
    return c

def cell(ws, row, col, value=None, fill=None, font=None, align=None, border=True, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:        c.fill   = fill
    if font:        c.font   = font or BODY_FONT
    if align:       c.alignment = align
    if border:      c.border = BORDER
    if num_format:  c.number_format = num_format
    return c

# ── Sheet 1: Goals Overview ────────────────────────────────────────────────────
def build_goals_overview(wb):
    ws = wb.create_sheet("Goals Overview")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="🎯  Personal Goal & Habit Tracker — Goals Overview")
    t.fill  = PatternFill("solid", fgColor="1A5276")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = ["Goal", "Category", "Priority", "Target Date",
               "Success Metric", "% Complete", "Status", "Notes"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 2, c, h)
    ws.row_dimensions[2].height = 22

    # Sample goals
    goals = [
        ("Run a half-marathon",          "Health",        "High",   date(2026, 10, 15),
         "Complete 13.1 mile race",         0.45, "On Track",  "Currently at 9 mile long run"),
        ("Save $5,000 emergency fund",   "Finance",       "High",   date(2026, 12, 31),
         "$5,000 in savings account",       0.62, "On Track",  "$3,100 saved so far"),
        ("Read 24 books this year",      "Learning",      "Medium", date(2026, 12, 31),
         "24 books logged in reading log", 0.50, "On Track",  "12 books completed"),
        ("Learn conversational Spanish", "Learning",      "Medium", date(2026, 9, 30),
         "Hold 10-min conversation",       0.30, "At Risk",   "Need to increase daily practice"),
        ("Meditate daily for 90 days",   "Health",        "Medium", date(2026, 9, 29),
         "90-day unbroken streak",         0.20, "At Risk",   "Missed 4 days in June"),
        ("Get promoted to Senior Dev",   "Career",        "High",   date(2026, 12, 31),
         "Offer letter received",          0.55, "On Track",  "Review meeting in September"),
        ("Declutter home — all rooms",   "Personal",      "Low",    date(2026, 8, 31),
         "Every room decluttered & clean", 0.80, "On Track",  "4 of 5 rooms done"),
        ("Reduce screen time < 2h/day",  "Health",        "Low",    date(2026, 7, 31),
         "Phone avg < 2 h for 30 days",   1.00, "Achieved",  "Sustained since May"),
        ("Complete online data course",  "Learning",      "Medium", date(2026, 11, 30),
         "Certificate received",           0.10, "Paused",    "Resuming after holiday"),
        ("Cook 3 new recipes per month", "Personal",      "Low",    date(2026, 12, 31),
         "36 new recipes across year",     0.42, "On Track",  "15 recipes so far"),
    ]

    status_fill = {
        "On Track": GREEN_FILL,
        "At Risk":  AMBER_FILL,
        "Achieved": BLUE_FILL,
        "Paused":   GREY_FILL,
    }

    for r, (goal, cat, pri, tgt, metric, pct, status, notes) in enumerate(goals, 3):
        row_fill = status_fill.get(status, PatternFill())
        data = [goal, cat, pri, tgt, metric, pct, status, notes]
        for c, val in enumerate(data, 1):
            cc = ws.cell(row=r, column=c, value=val)
            cc.border    = BORDER
            cc.font      = BODY_FONT
            cc.alignment = LEFT if c in (1, 5, 8) else CENTER
            if c in (1, 5, 8):
                cc.fill = INPUT_FILL
            elif c == 6:
                cc.fill = row_fill
                cc.number_format = "0%"
            elif c == 7:
                cc.fill = row_fill
            elif c == 4:
                cc.fill = INPUT_FILL
                cc.number_format = "DD-MMM-YYYY"
            else:
                cc.fill = INPUT_FILL
        ws.row_dimensions[r].height = 18

    # Column widths
    widths = [40, 14, 10, 14, 35, 12, 11, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Pie chart: Goals by category ──────────────────────────────────────────
    # Build a small summary table off to the right for the chart
    cats = {}
    for _, cat, *_ in goals:
        cats[cat] = cats.get(cat, 0) + 1

    ws.cell(row=2,  column=10, value="Category").font  = BOLD_FONT
    ws.cell(row=2,  column=11, value="Count").font     = BOLD_FONT
    for i, (cat, cnt) in enumerate(sorted(cats.items()), 3):
        ws.cell(row=i, column=10, value=cat)
        ws.cell(row=i, column=11, value=cnt)

    pie = PieChart()
    pie.title = "Goals by Category"
    pie.style = 10
    pie.dataLabels = None
    labels = Reference(ws, min_col=10, min_row=3, max_row=3 + len(cats) - 1)
    data   = Reference(ws, min_col=11, min_row=2, max_row=2 + len(cats))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width  = 14
    pie.height = 12
    ws.add_chart(pie, "J6")

    # ── Bar chart: % complete per goal ────────────────────────────────────────
    bar = BarChart()
    bar.type       = "bar"
    bar.title      = "Goal Completion %"
    bar.y_axis.title = "% Complete"
    bar.x_axis.title = "Goal"
    bar.style      = 10
    bar.grouping   = "clustered"

    goal_labels = Reference(ws, min_col=1, min_row=3, max_row=12)
    goal_pcts   = Reference(ws, min_col=6, min_row=2, max_row=12)
    bar.add_data(goal_pcts, titles_from_data=True)
    bar.set_categories(goal_labels)
    bar.width  = 20
    bar.height = 12
    ws.add_chart(bar, "J20")

    return ws


# ── Sheet 2: Daily Habits ──────────────────────────────────────────────────────
def build_daily_habits(wb):
    ws = wb.create_sheet("Daily Habits")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    year, month = 2026, 6
    days_in_month = 30
    today_day = 29  # June 29 2026

    # Title
    ws.merge_cells("A1:AF1")
    t = ws.cell(row=1, column=1, value="📅  Daily Habits — June 2026")
    t.fill  = PatternFill("solid", fgColor="1A5276")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Column headers: Habit | Icon | 1..31 | % Complete
    hdr(ws, 2, 1, "Habit")
    hdr(ws, 2, 2, "Target")
    for d in range(1, days_in_month + 1):
        day_date = date(year, month, d)
        day_label = f"{d}\n{day_date.strftime('%a')[:1]}"
        hdr(ws, 2, d + 2, day_label)
        ws.column_dimensions[get_column_letter(d + 2)].width = 4.2
    hdr(ws, 2, days_in_month + 3, "% Done")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions[get_column_letter(days_in_month + 3)].width = 9
    ws.row_dimensions[2].height = 28

    habits = [
        ("Exercise 30 min",    "Daily",   [1,1,0,1,1,1,0,1,1,1,1,0,1,1,1,0,1,1,1,1,0,1,1,0,1,1,1,1,0]),
        ("Read 20 minutes",    "Daily",   [1,1,1,1,1,0,1,1,1,1,1,1,1,0,1,1,1,1,1,1,1,0,1,1,1,1,1,1,1]),
        ("Meditate 10 min",    "Daily",   [1,0,1,1,1,1,0,0,1,1,1,0,1,1,0,1,1,1,0,1,1,1,0,1,1,0,1,1,1]),
        ("No alcohol",         "Daily",   [1,1,1,1,0,1,1,1,1,1,1,1,1,1,0,1,1,1,1,1,0,1,1,1,1,1,1,1,1]),
        ("Drink 64 oz water",  "Daily",   [1,1,0,1,1,1,1,1,0,1,1,1,1,1,1,1,1,0,1,1,1,1,1,1,0,1,1,1,1]),
        ("8 hours sleep",      "Daily",   [0,1,1,1,1,0,1,1,1,0,1,1,1,1,1,0,1,1,1,1,1,0,1,1,1,1,0,1,1]),
    ]

    habit_rows = {}
    for r_offset, (habit, target, completions) in enumerate(habits):
        r = r_offset + 3
        habit_rows[habit] = r

        ws.cell(row=r, column=1, value=habit).fill  = INPUT_FILL
        ws.cell(row=r, column=1).font               = BOLD_FONT
        ws.cell(row=r, column=1).alignment          = LEFT
        ws.cell(row=r, column=1).border             = BORDER

        ws.cell(row=r, column=2, value=target).fill  = INPUT_FILL
        ws.cell(row=r, column=2).font                = BODY_FONT
        ws.cell(row=r, column=2).alignment           = CENTER
        ws.cell(row=r, column=2).border              = BORDER

        done_count = 0
        for d in range(1, days_in_month + 1):
            col = d + 2
            if d <= today_day:
                val = completions[d - 1] if d - 1 < len(completions) else 0
                c = ws.cell(row=r, column=col, value=val)
                c.fill   = GREEN_FILL if val == 1 else RED_FILL
                c.font   = Font(name="Calibri", size=9, color="1E8449" if val else "C0392B")
                c.alignment = CENTER
                c.border = BORDER
                if val:
                    done_count += 1
            else:
                c = ws.cell(row=r, column=col, value="")
                c.fill   = PatternFill("solid", fgColor="F8F9FA")
                c.border = BORDER

        pct = done_count / today_day
        pc = ws.cell(row=r, column=days_in_month + 3, value=pct)
        pc.number_format = "0%"
        pc.font   = BOLD_FONT
        pc.fill   = GREEN_FILL if pct >= 0.8 else (AMBER_FILL if pct >= 0.6 else RED_FILL)
        pc.alignment = CENTER
        pc.border = BORDER
        ws.row_dimensions[r].height = 18

    # Daily score row
    score_row = len(habits) + 3
    ws.cell(row=score_row, column=1, value="Daily Score").fill  = HEADER_FILL
    ws.cell(row=score_row, column=1).font                       = WHITE_FONT
    ws.cell(row=score_row, column=1).alignment                  = CENTER
    ws.cell(row=score_row, column=1).border                     = BORDER
    ws.cell(row=score_row, column=2, value="").border           = BORDER

    for d in range(1, days_in_month + 1):
        col = d + 2
        if d <= today_day:
            daily_total = sum(
                (habits[h][2][d-1] if d-1 < len(habits[h][2]) else 0)
                for h in range(len(habits))
            )
            frac = daily_total / len(habits)
            c = ws.cell(row=score_row, column=col, value=f"{daily_total}/{len(habits)}")
            c.font   = Font(name="Calibri", bold=True, size=9,
                            color="1E8449" if frac >= 0.8 else ("E67E22" if frac >= 0.5 else "C0392B"))
            c.fill   = GREEN_FILL if frac >= 0.8 else (AMBER_FILL if frac >= 0.5 else RED_FILL)
            c.alignment = CENTER
            c.border = BORDER
        else:
            ws.cell(row=score_row, column=col).border = BORDER
    ws.row_dimensions[score_row].height = 18

    # Summary column score
    ws.cell(row=score_row, column=days_in_month + 3).border = BORDER

    # ── Bar chart: Habit completion % ─────────────────────────────────────────
    # Build helper table for chart
    chart_start_row = score_row + 2
    ws.cell(row=chart_start_row, column=1, value="Habit").font     = BOLD_FONT
    ws.cell(row=chart_start_row, column=2, value="% Complete").font = BOLD_FONT
    for i, (habit, _, completions) in enumerate(habits):
        r = chart_start_row + 1 + i
        done = sum(completions[:today_day])
        ws.cell(row=r, column=1, value=habit)
        ws.cell(row=r, column=2, value=done / today_day)
        ws.cell(row=r, column=2).number_format = "0%"

    bar = BarChart()
    bar.type        = "bar"
    bar.title       = "Habit Completion Rate — June 2026"
    bar.y_axis.title = "% of Days Completed"
    bar.style       = 10
    bar.grouping    = "clustered"
    bar.y_axis.numFmt = "0%"
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 1

    labels = Reference(ws, min_col=1, min_row=chart_start_row + 1,
                        max_row=chart_start_row + len(habits))
    data   = Reference(ws, min_col=2, min_row=chart_start_row,
                        max_row=chart_start_row + len(habits))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.width  = 22
    bar.height = 14
    ws.add_chart(bar, "C11")

    return ws


# ── Sheet 3: Weekly Review ─────────────────────────────────────────────────────
def build_weekly_review(wb):
    ws = wb.create_sheet("Weekly Review")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="📝  Weekly Review — June 2026")
    t.fill  = PatternFill("solid", fgColor="1A5276")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16

    # 4 completed weeks + partial week 5
    weeks = [
        {
            "label":    "Week 1 — 1 Jun to 7 Jun",
            "well":     "Maintained exercise streak all 7 days. Finished two book chapters ahead of schedule.",
            "challenge":"Struggled with meditation after late work nights Tuesday and Wednesday.",
            "kept":     "Exercise, Reading, No Alcohol, Water intake",
            "missed":   "Meditation (2 days). Late nights disrupted routine.",
            "focus":    "Set a phone alarm at 9 pm as a meditation reminder.",
            "scores":   [5, 4, 3, 4, 5, 4, 4, 3, 5, 4],  # 10 goal scores
        },
        {
            "label":    "Week 2 — 8 Jun to 14 Jun",
            "well":     "Completed first 6-mile run without stopping — biggest distance yet.",
            "challenge":"Spanish practice fell to zero; kept de-prioritising it.",
            "kept":     "Exercise, Reading, Water, Sleep",
            "missed":   "Meditation (1 day), Spanish study (5 days).",
            "focus":    "Block 20 min Spanish practice every weekday morning.",
            "scores":   [5, 4, 4, 2, 4, 5, 5, 3, 3, 4],
        },
        {
            "label":    "Week 3 — 15 Jun to 21 Jun",
            "well":     "Hit savings milestone: $3,000 saved. Decluttered the spare room.",
            "challenge":"Thursday social event caused one missed no-alcohol day.",
            "kept":     "Exercise, Reading, Meditation, Water, Sleep",
            "missed":   "No Alcohol (1 day — planned social event).",
            "focus":    "Plan for social events in advance — decide on limits beforehand.",
            "scores":   [5, 5, 3, 3, 5, 5, 5, 3, 3, 5],
        },
        {
            "label":    "Week 4 — 22 Jun to 28 Jun",
            "well":     "Spanish momentum returning: used the app 5 of 7 days. Sleep improving.",
            "challenge":"Exercise skipped Saturday due to injury scare (left knee). Resting.",
            "kept":     "Reading, Meditation, No Alcohol, Water, Sleep",
            "missed":   "Exercise (1 day — knee rest). Monitoring closely.",
            "focus":    "Substitute swimming for running while knee recovers. Book physio appt.",
            "scores":   [4, 5, 4, 4, 4, 5, 5, 4, 4, 4],
        },
    ]

    prompts = [
        ("What went well this week?",         "well"),
        ("What was challenging?",             "challenge"),
        ("Which habits did I maintain?",      "kept"),
        ("Which habits did I miss, and why?", "missed"),
        ("What is my focus for next week?",   "focus"),
    ]

    cur_row = 2
    for week in weeks:
        # Week header
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        wh = ws.cell(row=cur_row, column=1, value=week["label"])
        wh.fill  = SUBHDR_FILL
        wh.font  = WHITE_FONT
        wh.alignment = LEFT
        wh.border    = BORDER
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Prompt rows
        for label, key in prompts:
            ws.cell(row=cur_row, column=1, value=label).font      = BOLD_FONT
            ws.cell(row=cur_row, column=1).alignment              = LEFT
            ws.cell(row=cur_row, column=1).border                 = BORDER
            ws.cell(row=cur_row, column=1).fill                   = PatternFill("solid", fgColor="EBF5FB")
            ws.merge_cells(f"B{cur_row}:D{cur_row}")
            vc = ws.cell(row=cur_row, column=2, value=week[key])
            vc.fill      = INPUT_FILL
            vc.font      = BODY_FONT
            vc.alignment = LEFT
            vc.border    = BORDER
            ws.row_dimensions[cur_row].height = 30
            cur_row += 1

        # Goal momentum scores
        ws.cell(row=cur_row, column=1, value="Goal momentum (1–5)").font   = BOLD_FONT
        ws.cell(row=cur_row, column=1).alignment                           = LEFT
        ws.cell(row=cur_row, column=1).border                              = BORDER
        ws.cell(row=cur_row, column=1).fill                                = PatternFill("solid", fgColor="EBF5FB")
        goal_names_short = ["Half-marathon","Emergency fund","Read 24 books","Spanish",
                             "Meditate 90d","Promotion","Declutter","Screen time","Data course","Cooking"]
        for i, (name, score) in enumerate(zip(goal_names_short, week["scores"])):
            col = i + 2 if i < 3 else None
            # Put all scores in merged cell as text
        score_text = " | ".join(f"{n}: {s}" for n, s in zip(goal_names_short, week["scores"]))
        ws.merge_cells(f"B{cur_row}:D{cur_row}")
        sc = ws.cell(row=cur_row, column=2, value=score_text)
        sc.fill      = INPUT_FILL
        sc.font      = BODY_FONT
        sc.alignment = LEFT
        sc.border    = BORDER
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        # Spacer
        ws.row_dimensions[cur_row].height = 8
        cur_row += 1

    # ── Line chart: weekly momentum scores per goal ────────────────────────────
    # Build helper data table
    chart_row = cur_row + 1
    ws.cell(row=chart_row, column=1, value="Week").font = BOLD_FONT
    goal_names_short = ["Half-marathon","Savings","Reading","Spanish",
                        "Meditate","Promotion","Declutter","Screen","Data Crs","Cooking"]
    for i, name in enumerate(goal_names_short):
        ws.cell(row=chart_row, column=i + 2, value=name).font = BOLD_FONT

    for w_i, week in enumerate(weeks):
        r = chart_row + 1 + w_i
        ws.cell(row=r, column=1, value=f"Week {w_i+1}")
        for i, score in enumerate(week["scores"]):
            ws.cell(row=r, column=i + 2, value=score)

    line = LineChart()
    line.title        = "Weekly Goal Momentum (1–5)"
    line.y_axis.title = "Momentum Score"
    line.x_axis.title = "Week"
    line.style        = 10
    line.y_axis.scaling.min = 1
    line.y_axis.scaling.max = 5

    data = Reference(ws, min_col=2, max_col=11,
                     min_row=chart_row, max_row=chart_row + len(weeks))
    cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(weeks))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.width  = 26
    line.height = 14
    ws.add_chart(line, "A" + str(cur_row + 2))

    return ws


# ── Sheet 4: Milestones ────────────────────────────────────────────────────────
def build_milestones(wb):
    ws = wb.create_sheet("Milestones")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="🏆  Milestones & Achievements — 2026")
    t.fill  = PatternFill("solid", fgColor="1A5276")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for c, h in enumerate(["Date", "Milestone", "Related Goal", "How I Felt"], 1):
        hdr(ws, 2, c, h)
    ws.row_dimensions[2].height = 22

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 50

    milestones = [
        (date(2026, 1, 5),  "Set up this tracker and committed to 10 clear goals for 2026",
         "All goals",       "Excited and a little nervous. Felt like this time it's different."),
        (date(2026, 2, 12), "Read my 4th book of the year — on track for 24",
         "Read 24 books",   "Satisfied. The reading habit has genuinely stuck this time."),
        (date(2026, 3, 1),  "First month with no missed exercise days — 28/28 streak",
         "Half-marathon",   "Proud. Showed myself I can be consistent when I plan properly."),
        (date(2026, 3, 22), "Passed the $1,000 emergency fund threshold",
         "Emergency fund",  "Relieved. Finally building real financial security."),
        (date(2026, 4, 7),  "Completed 30-day meditation streak",
         "Meditate 90 days","Calm and focused. The habit is starting to feel effortless."),
        (date(2026, 4, 19), "Ran 6 miles for the first time — race-pace finish",
         "Half-marathon",   "Euphoric. Couldn't believe how far I'd come in 4 months."),
        (date(2026, 5, 3),  "Decluttered the kitchen and living room — two rooms done",
         "Declutter home",  "Light. The space feels totally different now."),
        (date(2026, 5, 20), "Screen time app showed first week averaging under 2 hours",
         "Screen time",     "Surprised it worked. More evening time feels like a gift."),
        (date(2026, 6, 1),  "Hit $3,000 in savings — 60% of emergency fund goal",
         "Emergency fund",  "Motivated. At this rate I'll hit $5k by October."),
        (date(2026, 6, 10), "Read book #12 — halfway to the 24-book challenge",
         "Read 24 books",   "Delighted. The evening reading habit is now non-negotiable."),
        (date(2026, 6, 15), "Best 6-mile time: 51 min 34 sec — new PR",
         "Half-marathon",   "Emotional. Months of early mornings paid off."),
        (date(2026, 6, 21), "Decluttered spare room — 4 of 5 rooms complete",
         "Declutter home",  "Accomplished. One room to go — the garage."),
        (date(2026, 6, 28), "First Spanish conversation with a native speaker (10 min!)",
         "Spanish",         "Giddy. All those awkward practice sessions were worth it."),
    ]

    row_fills = [GREEN_FILL, AMBER_FILL, BLUE_FILL]
    for r_offset, (dt, milestone, goal, feeling) in enumerate(milestones):
        r   = r_offset + 3
        fill = GREEN_FILL
        for c, val in enumerate([dt, milestone, goal, feeling], 1):
            cc = ws.cell(row=r, column=c, value=val)
            cc.fill      = fill
            cc.font      = BODY_FONT
            cc.alignment = LEFT if c > 1 else CENTER
            cc.border    = BORDER
            if c == 1:
                cc.number_format = "DD-MMM-YYYY"
        ws.row_dimensions[r].height = 20

    return ws


# ── Sheet 5: Year at a Glance ──────────────────────────────────────────────────
def build_year_at_glance(wb):
    ws = wb.create_sheet("Year at a Glance")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    ws.merge_cells("A1:BD1")
    t = ws.cell(row=1, column=1, value="📊  Year at a Glance — Habit Heatmap 2026  (darker = more consistent)")
    t.fill  = PatternFill("solid", fgColor="1A5276")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    habits = [
        "Exercise 30 min",
        "Read 20 minutes",
        "Meditate 10 min",
        "No alcohol",
        "Drink 64 oz water",
        "8 hours sleep",
    ]

    hdr(ws, 2, 1, "Habit")
    hdr(ws, 2, 2, "Month")
    for w in range(1, 53):
        hdr(ws, 2, w + 2, f"W{w}", font=Font(name="Calibri", bold=True, color="FFFFFF", size=8))
        ws.column_dimensions[get_column_letter(w + 2)].width = 3.5

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.row_dimensions[2].height = 22

    # Simulate realistic weekly scores (0-7) for each habit across 52 weeks
    import random
    random.seed(42)

    # Scores ramp up over the year — early weeks lower, recent weeks higher
    def habit_week_score(habit_idx, week):
        """Generate a plausible completion score for a given habit and week."""
        # Base adherence improves over the year
        base = 3.5 + (week / 52) * 2.0
        # Each habit has slightly different volatility
        noise = random.gauss(0, 1.2)
        score = base + noise
        # Some habits are harder (meditation, Spanish → not tracked here)
        if habit_idx in (2,):  # meditation is harder
            score -= 0.8
        # Future weeks after today (week 27 = end of June) are blank
        if week > 26:
            return None
        return max(0, min(7, round(score)))

    fills_by_score = {
        0: PatternFill("solid", fgColor="FDFEFE"),   # white
        1: PatternFill("solid", fgColor="D5F5E3"),   # very light green
        2: PatternFill("solid", fgColor="ABEBC6"),
        3: PatternFill("solid", fgColor="82E0AA"),
        4: PatternFill("solid", fgColor="52BE80"),
        5: PatternFill("solid", fgColor="27AE60"),
        6: PatternFill("solid", fgColor="1E8449"),
        7: PatternFill("solid", fgColor="145A32"),   # dark green
    }
    score_font = {
        **{k: Font(name="Calibri", size=8, color="2ECC71") for k in range(4)},
        **{k: Font(name="Calibri", size=8, color="FFFFFF") for k in range(4, 8)},
    }

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    week_month = {w: months[min(int((w - 1) * 12 / 52), 11)] for w in range(1, 53)}

    for h_idx, habit in enumerate(habits):
        r = h_idx + 3
        ws.cell(row=r, column=1, value=habit).font      = BOLD_FONT
        ws.cell(row=r, column=1).fill                   = INPUT_FILL
        ws.cell(row=r, column=1).alignment              = LEFT
        ws.cell(row=r, column=1).border                 = BORDER
        ws.cell(row=r, column=2, value=week_month.get(1, "Jan")).font = SMALL_FONT
        ws.cell(row=r, column=2).alignment              = CENTER
        ws.cell(row=r, column=2).border                 = BORDER

        for w in range(1, 53):
            col = w + 2
            score = habit_week_score(h_idx, w)
            if score is None:
                c = ws.cell(row=r, column=col, value="")
                c.fill   = PatternFill("solid", fgColor="F2F3F4")
                c.border = BORDER
            else:
                c = ws.cell(row=r, column=col, value=score)
                c.fill      = fills_by_score.get(score, fills_by_score[0])
                c.font      = score_font.get(score, BODY_FONT)
                c.alignment = CENTER
                c.border    = BORDER
        ws.row_dimensions[r].height = 20

    # Month label row
    month_row = len(habits) + 3
    ws.merge_cells(f"A{month_row}:B{month_row}")
    ws.cell(row=month_row, column=1, value="Month →").font = SMALL_FONT
    ws.cell(row=month_row, column=1).alignment = CENTER

    prev_month = None
    for w in range(1, 53):
        m = week_month[w]
        col = w + 2
        if m != prev_month:
            ws.cell(row=month_row, column=col, value=m).font = Font(name="Calibri", bold=True, size=8)
            ws.cell(row=month_row, column=col).alignment = CENTER
            prev_month = m

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = month_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = BOLD_FONT
    labels = ["0 days","1 day","2 days","3 days","4 days","5 days","6 days","7 days"]
    for i, label in enumerate(labels):
        col = i + 2
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill   = fills_by_score[i]
        c.font   = Font(name="Calibri", size=8,
                        color="2ECC71" if i < 4 else "FFFFFF")
        c.alignment = CENTER
        c.border = BORDER
        if col > 2:
            ws.column_dimensions[get_column_letter(col)].width = 3.5

    # ── Annual summary bar chart ───────────────────────────────────────────────
    # Build summary: average weekly score per habit across completed weeks
    sum_row = legend_row + 3
    ws.cell(row=sum_row, column=1, value="Habit").font = BOLD_FONT
    ws.cell(row=sum_row, column=2, value="Avg Days/Week").font = BOLD_FONT

    import random
    random.seed(42)
    for h_idx, habit in enumerate(habits):
        r = sum_row + 1 + h_idx
        scores = [habit_week_score(h_idx, w) for w in range(1, 27)]
        scores = [s for s in scores if s is not None]
        avg = sum(scores) / len(scores) if scores else 0
        ws.cell(row=r, column=1, value=habit)
        ws.cell(row=r, column=2, value=round(avg, 1))

    bar = BarChart()
    bar.type        = "bar"
    bar.title       = "Average Habit Days per Week (YTD)"
    bar.y_axis.title = "Days / Week (max 7)"
    bar.style       = 10
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 7

    labels_ref = Reference(ws, min_col=1, min_row=sum_row + 1,
                            max_row=sum_row + len(habits))
    data_ref   = Reference(ws, min_col=2, min_row=sum_row,
                            max_row=sum_row + len(habits))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(labels_ref)
    bar.width  = 20
    bar.height = 14
    ws.add_chart(bar, f"C{legend_row + 3}")

    return ws


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet

    build_goals_overview(wb)
    build_daily_habits(wb)
    build_weekly_review(wb)
    build_milestones(wb)
    build_year_at_glance(wb)

    out_dir  = "/home/davidj/Projects/calc_personal_goal_habit_tracker/personal-goal-habit-tracker"
    xlsx_out = f"{out_dir}/personal-goal-habit-tracker.xlsx"
    ots_out  = f"{out_dir}/personal-goal-habit-tracker.ots"

    wb.save(xlsx_out)
    print(f"Saved xlsx: {xlsx_out}")

    # Convert to OTS via LibreOffice headless
    result = subprocess.run(
        [
            "libreoffice", "--headless", "--convert-to", "ots",
            "--outdir", out_dir, xlsx_out
        ],
        capture_output=True, text=True, timeout=60
    )
    print(result.stdout)
    if result.returncode != 0:
        print("STDERR:", result.stderr)
    else:
        print(f"Saved ots: {ots_out}")


if __name__ == "__main__":
    main()
